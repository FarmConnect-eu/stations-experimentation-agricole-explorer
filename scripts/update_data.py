#!/usr/bin/env python3
"""
Mise a jour de la base consolidee des stations d'experimentation agricole.

Ce script :
1. Telecharge la derniere version de la carte APCA depuis Grist (data.gouv ANCT)
2. Fusionne avec les fermes complementaires (Digifermes/F@rmXP/IRFEL non listees APCA)
3. Genere le fichier consolide data/stations_experimentation_consolide.xlsx
4. Le workflow GitHub Actions s'occupera ensuite de regenerer index.html

Usage local :
    pip install openpyxl requests
    python3 scripts/update_data.py

Source APCA officielle :
    https://grist.incubateur.anct.gouv.fr (doc 7zGLWG4fUHaJaNj9robyBS,
    table VF_Carto_Stations_Expes_25_02_2026)
"""
import sys
import json
import unicodedata
import re
from pathlib import Path

try:
    import requests
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"ERREUR : module manquant ({e}). Lance :\n    pip install openpyxl requests", file=sys.stderr)
    sys.exit(1)

# =====================================================================
# CONFIGURATION
# =====================================================================

ROOT = Path(__file__).parent.parent
OUTPUT_XLSX = ROOT / "data" / "stations_experimentation_consolide.xlsx"

# Source officielle Grist APCA (carte uMap des Chambres d'agriculture)
APCA_GRIST_DOC = "7zGLWG4fUHaJaNj9robyBS"
APCA_GRIST_TABLE = "VF_Carto_Stations_Expes_25_02_2026"
APCA_XLSX_URL = f"https://grist.incubateur.anct.gouv.fr/api/docs/{APCA_GRIST_DOC}/download/xlsx?tableId={APCA_GRIST_TABLE}"

HEADERS = [
    "Nom", "Region", "Contact", "Tel", "Mail", "Type de production",
    "Sujets d'experimentation", "Filiere", "Especes etudiees",
    "Reseaux de fermes experimentales", "Adresse postale",
    "latitude", "longitude", "Site internet",
]

# =====================================================================
# FERMES COMPLEMENTAIRES (presentes dans Digifermes/F@rmXP/IRFEL mais
# ABSENTES de la carte APCA). Ces 15 lignes sont ajoutees apres la fusion.
# Schema identique aux 14 colonnes du template.
# A maintenir manuellement si de nouveaux reseaux/fermes apparaissent.
# =====================================================================

FERMES_COMPLEMENTAIRES = [
    {
        "nom": "DIGIFERME ARMEFLHOR",
        "region": "La Reunion",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Outils numeriques, Veille technologique, Vulgarisation",
        "filiere": "Fruits et legumes",
        "especes": "Maraichage sous abri, Arboriculture fruitiere, PPAM, Horticulture",
        "reseaux": "Reseau Digifermes",
        "adresse": "1 chemin de l'IRFA, Bassin Martin, 97410 Saint-Pierre",
        "lat": -21.3393, "lon": 55.4781,
        "site": "https://armeflhor.fr/",
    },
    {
        "nom": "DIGIFERME CIRI (IFIP)",
        "region": "Bretagne",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Nutrition, Sante, Environnement, Bien-etre animal, Reproduction, Batiments-equipements",
        "filiere": "Elevage (porc)",
        "especes": "Porcs",
        "reseaux": "Reseau Digifermes",
        "adresse": "La Motte au Vicomte, 35850 Romille",
        "lat": 48.2150, "lon": -1.8736,
        "site": "https://ifip.asso.fr/",
    },
    {
        "nom": "DIGIFERME de Boigneville",
        "region": "Ile-de-France",
        "contact": "", "tel": "", "mail": "",
        "type": "Mixte AB et conventionnel",
        "sujets": "Zero saisie, Stations meteo connectees, Capteurs, OAD, Robots, Gestion des adventices, Irrigation de precision",
        "filiere": "Grandes cultures",
        "especes": "Ble, Orge, Mais, Colza, Pois, Chanvre",
        "reseaux": "Reseau Digifermes",
        "adresse": "Station experimentale ARVALIS, 91720 Boigneville",
        "lat": 48.3275, "lon": 2.3850,
        "site": "https://www.arvalis.fr/l-institut/nous-connaitre/stations-laboratoires-arvalis/station-boigneville",
    },
    {
        "nom": "DIGIFERME de Chamberet",
        "region": "Nouvelle-Aquitaine",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Tracabilite zootechnique et sanitaire, Recherche appliquee, Formation",
        "filiere": "Elevage (equin)",
        "especes": "Equins",
        "reseaux": "Reseau Digifermes",
        "adresse": "Plateau technique IFCE, 19370 Chamberet",
        "lat": 45.5872, "lon": 1.7194,
        "site": "https://www.ifce.fr/ifce/recherche/plateaux-techniques/chamberet/",
    },
    {
        "nom": "DIGIFERME La Cazotte",
        "region": "Occitanie",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Detection precoce sanitaire, Bien-etre animal, Conditions de travail, Pesage automatique, Suivi pousse herbe",
        "filiere": "Elevage (ovin)",
        "especes": "Ovins lait (AOP Roquefort), Ovins viande, Bovins",
        "reseaux": "Reseau Digifermes",
        "adresse": "Route de Bournac, 12400 Saint-Affrique",
        "lat": 43.9586, "lon": 2.8861,
        "site": "https://la-cazotte.educagri.fr/",
    },
    {
        "nom": "DIGIFERME le CIRVEAU",
        "region": "Bretagne",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Batiment d'elevage du futur, Automatisation, Podometres, Antennes UHF, Innovation technologique",
        "filiere": "Elevage (bovin)",
        "especes": "Veaux de boucherie",
        "reseaux": "Reseau Digifermes",
        "adresse": "Le Rheu, 56430 Mauron",
        "lat": 48.0867, "lon": -2.2861,
        "site": "https://digifermes.com/blog/axes-de-travail/mauron-56/",
    },
    {
        "nom": "DIGIFERME de Plumecoq",
        "region": "Grand Est",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Robotique, Alternatives au desherbage chimique, Tracabilite, IA, OAD, Automatisation",
        "filiere": "Viticulture",
        "especes": "Vigne (Champagne)",
        "reseaux": "Reseau Digifermes",
        "adresse": "51530 Chouilly",
        "lat": 49.0653, "lon": 4.0019,
        "site": "https://digifermes.com/blog/axes-de-travail/plumecoq-51/",
    },
    {
        "nom": "DIGIFERME de Saint-Hilaire-en-Woevre",
        "region": "Grand Est",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Desherbage drone, Robotique, Gestion paturage, Autonomie proteique, Bien-etre animal, Reduction penibilite",
        "filiere": "Elevage (bovin) et grandes cultures",
        "especes": "Vaches allaitantes, Prairies, Mais",
        "reseaux": "Reseau Digifermes",
        "adresse": "16 Rue du Moulin de Moncelle, 55160 Saint-Hilaire-en-Woevre",
        "lat": 48.9914, "lon": 5.7197,
        "site": "https://www.arvalis.fr/l-institut/nous-connaitre/stations-laboratoires-arvalis/ferme-experimentale-lorraine",
    },
    {
        "nom": "DIGIFERME V'innopole Sud-Ouest",
        "region": "Occitanie",
        "contact": "", "tel": "", "mail": "",
        "type": "Conventionnel",
        "sujets": "Pilotage numerique, Robotique, Capteurs sol, Pulverisateur innovant, Imagerie radar, Oenologie",
        "filiere": "Viticulture",
        "especes": "Vigne (300 cepages)",
        "reseaux": "Reseau Digifermes",
        "adresse": "Brames-Aigues, 81310 Lisle-sur-Tarn",
        "lat": 43.8567, "lon": 1.8092,
        "site": "https://www.vignevin-occitanie.com/qui-sommes-nous/pole-sud-ouest/",
    },
    {
        "nom": "ANPN",
        "region": "Nouvelle-Aquitaine",
        "contact": "Maud THOMAS", "tel": "05 53 01 60 08", "mail": "contact@anpn.eu",
        "type": "Mixte AB et conventionnel",
        "sujets": "Evaluation varietale, Gestion enherbement, Maladies, Ravageurs (entomologie, biocontrole), Fertilisation, Irrigation, Biodiversite",
        "filiere": "Arboriculture fruitiere",
        "especes": "Noisette, Noix, Amande",
        "reseaux": "Reseau IRFEL",
        "adresse": "Lieu-dit Louberie, 47290 Cancon",
        "lat": 44.5275, "lon": 0.6311,
        "site": "https://www.anpn.eu/",
    },
    {
        "nom": "APEF",
        "region": "Hauts-de-France",
        "contact": "Pierre VARLET", "tel": "03 21 07 89 89", "mail": "contact@endive.fr",
        "type": "Conventionnel",
        "sujets": "Gestion enherbement, Irrigation, Maladies, Ravageurs, Evaluation et creation varietale, Emballages loi AGEC",
        "filiere": "Fruits et legumes",
        "especes": "Endive (et produits derives Carmine)",
        "reseaux": "Reseau IRFEL",
        "adresse": "2 Rue des Fleurs, 62000 Arras",
        "lat": 50.2914, "lon": 2.7772,
        "site": "https://www.endive.fr/",
    },
    {
        "nom": "AREFE",
        "region": "Grand Est",
        "contact": "", "tel": "03 29 89 58 18", "mail": "arefe@orange.fr",
        "type": "Mixte AB et conventionnel",
        "sujets": "Evaluation varietale, Gestion enherbement, Maladies, Ravageurs, Fertilisation, Durabilite des sols, Biodiversite",
        "filiere": "Arboriculture fruitiere",
        "especes": "Mirabelle, Prune, Cerise acide",
        "reseaux": "Reseau IRFEL",
        "adresse": "Rue Arnay Le Duc, Hattonville, 55210 Vigneulles-les-Hattonchatel",
        "lat": 48.9803, "lon": 5.7044,
        "site": "http://www.prunes-et-mirabelles-de-lorraine.com/",
    },
    {
        "nom": "ARELPAL",
        "region": "Pays de la Loire",
        "contact": "Sylvain GERARD", "tel": "02 28 27 03 00", "mail": "sylvain.gerard@arelpal.org",
        "type": "Mixte AB et conventionnel",
        "sujets": "Evaluation varietale, Maladies et ravageurs, Irrigation, Gestion enherbement, Technique culturale",
        "filiere": "Fruits et legumes",
        "especes": "Laitue, Jeunes pousses, Poireau, Radis, Tomate, Concombre",
        "reseaux": "Reseau IRFEL",
        "adresse": "La Metairie Neuve, 44860 Pont-Saint-Martin",
        "lat": 47.1278, "lon": -1.5783,
        "site": "https://www.arelpal.org/",
    },
    {
        "nom": "Station experimentale de Creysse",
        "region": "Nouvelle-Aquitaine",
        "contact": "", "tel": "05 65 32 22 22", "mail": "contact@stationcreysse.fr",
        "type": "Mixte AB et conventionnel",
        "sujets": "Evaluation varietale, Conduite culturale, Maladies, Ravageurs, Conservation post-recolte",
        "filiere": "Arboriculture fruitiere",
        "especes": "Noix",
        "reseaux": "Reseau IRFEL",
        "adresse": "30 Route de Perrical, 46600 Creysse",
        "lat": 44.8889, "lon": 1.6017,
        "site": "https://www.stationexperimentaledecreysse.fr/",
    },
    {
        "nom": "GRAB",
        "region": "Provence-Alpes-Cote d'Azur",
        "contact": "", "tel": "04 90 84 01 70", "mail": "secretariat@grab.fr",
        "type": "Agriculture biologique",
        "sujets": "Agriculture biologique, Evaluation varietale, Gestion enherbement, Couverts vegetaux, Biocontrole, Bandes fleuries, Barrieres physiques",
        "filiere": "Fruits et legumes, Arboriculture fruitiere",
        "especes": "Pomme, Poire, Abricot, Peche, Cerise, Amande, Salade, Tomate, Aubergine, Courgette",
        "reseaux": "Reseau IRFEL",
        "adresse": "Maison de la Bio, 255 chemin de la Castelette, 84911 Avignon",
        "lat": 43.9493, "lon": 4.8055,
        "site": "https://www.grab.fr/",
    },
]


# =====================================================================
# ETAPES
# =====================================================================

def step(n, label):
    print(f"\n=== Etape {n} : {label} ===")

def clean_cell(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace("\u2014", "-").replace("\u2013", "-")  # em-dash, en-dash
    return s.strip()


def fetch_apca_xlsx():
    """Telecharge le xlsx APCA depuis Grist."""
    step(1, "Telechargement APCA Grist")
    print(f"URL : {APCA_XLSX_URL}")
    r = requests.get(APCA_XLSX_URL, timeout=30)
    r.raise_for_status()
    tmp = ROOT / "data" / ".apca_raw.xlsx"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_bytes(r.content)
    print(f"OK : {len(r.content)} octets telecharges dans {tmp}")
    return tmp


def load_apca_rows(xlsx_path):
    """Lit le xlsx APCA et retourne une liste de dicts."""
    step(2, "Lecture du xlsx APCA")
    ws = load_workbook(xlsx_path, data_only=True).active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        row = {
            "nom": clean_cell(r[0]),
            "region": clean_cell(r[1]),
            "contact": clean_cell(r[2]),
            "tel": clean_cell(r[3]),
            "mail": clean_cell(r[4]),
            "type": clean_cell(r[5]),
            "sujets": clean_cell(r[6]),
            "filiere": clean_cell(r[7]),
            "especes": clean_cell(r[8]),
            "reseaux": clean_cell(r[9]) or "Reseau Chambre d'agriculture",
            "adresse": clean_cell(r[10]),
            "lat": r[11] if isinstance(r[11], (int, float)) else (float(str(r[11]).strip()) if r[11] and str(r[11]).strip() else None),
            "lon": r[12] if isinstance(r[12], (int, float)) else (float(str(r[12]).strip()) if r[12] and str(r[12]).strip() else None),
            "site": clean_cell(r[13]),
        }
        # Harmonisation : "Aucun" devient le reseau Chambre d'agriculture
        if row["reseaux"] == "Aucun":
            row["reseaux"] = "Reseau Chambre d'agriculture"
        # Harmonisation : "AB" -> "Agriculture biologique"
        if row["type"] == "AB":
            row["type"] = "Agriculture biologique"
        rows.append(row)
    print(f"OK : {len(rows)} stations APCA lues")
    return rows


def merge(apca_rows):
    """Fusionne APCA + fermes complementaires."""
    step(3, "Fusion APCA + fermes complementaires")
    # Les fermes complementaires sont definies comme ABSENTES d'APCA par construction
    # (liste maintenue manuellement). On ajoute simplement les 15 lignes.
    final = list(apca_rows)
    for f in FERMES_COMPLEMENTAIRES:
        final.append(dict(f))
    print(f"OK : {len(final)} lignes finales ({len(apca_rows)} APCA + {len(FERMES_COMPLEMENTAIRES)} complementaires)")
    return final


def write_xlsx(rows, output_path):
    """Ecrit le xlsx final avec mise en forme."""
    step(4, "Generation du xlsx consolide")
    wb = Workbook()
    ws = wb.active
    ws.title = "Stations experimentales"

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    widths = [42, 22, 24, 16, 35, 24, 55, 26, 50, 28, 50, 11, 11, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, row in enumerate(rows, start=2):
        values = [
            row["nom"], row["region"], row["contact"], row["tel"], row["mail"],
            row["type"], row["sujets"], row["filiere"], row["especes"],
            row["reseaux"], row["adresse"],
            row["lat"] if row["lat"] is not None else "",
            row["lon"] if row["lon"] is not None else "",
            row["site"],
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = data_font
            c.alignment = data_align
            c.border = border
            if col in (12, 13) and isinstance(val, (int, float)):
                c.number_format = "0.000000"

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"OK : {output_path} ecrit ({ws.max_row - 1} lignes + 1 en-tete)")


def stats(rows):
    """Affiche la repartition par reseau et par type."""
    step(5, "Statistiques")
    from collections import Counter
    res = Counter(r["reseaux"] for r in rows)
    typ = Counter(r["type"] for r in rows if r["type"])
    print("\nReseaux :")
    for k, v in res.most_common():
        print(f"  {v:3d}  {k}")
    print("\nTypes :")
    for k, v in typ.most_common():
        print(f"  {v:3d}  {k}")


def main():
    apca_xlsx = fetch_apca_xlsx()
    apca_rows = load_apca_rows(apca_xlsx)
    final_rows = merge(apca_rows)
    write_xlsx(final_rows, OUTPUT_XLSX)
    stats(final_rows)
    print(f"\nFini. Fichier consolide : {OUTPUT_XLSX}")
    print("\nProchaine etape :")
    print("  git add data/stations_experimentation_consolide.xlsx")
    print("  git commit -m 'Maj donnees stations'")
    print("  git push")
    print("Le workflow GitHub Actions regenerera ensuite index.html automatiquement.")


if __name__ == "__main__":
    main()
