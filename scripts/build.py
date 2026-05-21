#!/usr/bin/env python3
"""
Regenere index.html depuis data/stations_experimentation_consolide.xlsx.

Usage local :
    python3 scripts/build.py

Usage CI : appele automatiquement par .github/workflows/build.yml
quand data/*.xlsx est modifie sur main.

Dependances : openpyxl (pip install openpyxl)
"""
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERREUR : openpyxl n'est pas installe. Lance : pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "data" / "stations_experimentation_consolide.xlsx"
TEMPLATE = ROOT / "scripts" / "template.html"
OUT = ROOT / "index.html"

if not XLSX.exists():
    print(f"ERREUR : {XLSX} introuvable", file=sys.stderr)
    sys.exit(1)
if not TEMPLATE.exists():
    print(f"ERREUR : {TEMPLATE} introuvable", file=sys.stderr)
    sys.exit(1)

# Lire les stations depuis le xlsx
ws = load_workbook(XLSX, data_only=True).active
stations = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    s = {
        "nom": row[0] or "",
        "region": row[1] or "",
        "contact": row[2] or "",
        "tel": row[3] or "",
        "mail": row[4] or "",
        "type": row[5] or "",
        "sujets": row[6] or "",
        "filiere": row[7] or "",
        "especes": row[8] or "",
        "reseaux": row[9] or "Reseau Chambre d'agriculture",
        "adresse": row[10] or "",
        "lat": row[11] if isinstance(row[11], (int, float)) else None,
        "lon": row[12] if isinstance(row[12], (int, float)) else None,
        "site": row[13] or "",
    }
    stations.append(s)

data_json = json.dumps(stations, ensure_ascii=False, separators=(',', ':'))

# Charger le template et injecter les donnees
html = TEMPLATE.read_text(encoding="utf-8")
if "__DATA__" not in html:
    print("ERREUR : le template ne contient pas le placeholder __DATA__", file=sys.stderr)
    sys.exit(1)
html = html.replace("__DATA__", data_json)

OUT.write_text(html, encoding="utf-8")
print(f"OK : {OUT} regenere ({len(stations)} stations, {len(html)} caracteres)")
